#!/usr/bin/env python3
"""Generate deterministic fictional fixtures for BrewGo Day2 Live Tasks."""
from __future__ import annotations

import html
import json
import random
from collections import defaultdict
from datetime import date, timedelta
from pathlib import Path

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.worksheet.table import Table, TableStyleInfo

ROOT = Path(__file__).resolve().parents[1]
DAY2_ROOT = ROOT / "classroom" / "day2-live-tasks"
AMAZON_ROOT = DAY2_ROOT / "01-amazon-competitor-discovery" / "input" / "offline"
INSTAGRAM_ROOT = DAY2_ROOT / "02-instagram-lead-discovery" / "input" / "offline"
DATA_PATH = DAY2_ROOT / "03-data-analysis-dashboard" / "input" / "business-performance.xlsx"
REFERENCE_DASHBOARD = ROOT / "instructor" / "day2-data-dashboard-reference.html"
REFERENCE_METRICS = ROOT / "instructor" / "day2-data-dashboard-reference-metrics.json"
LABEL = "FICTIONAL / TEACHING SNAPSHOT"

CSS = """
:root{font-family:Inter,ui-sans-serif,system-ui,-apple-system,BlinkMacSystemFont,"Segoe UI",sans-serif;color:#17221b;background:#f4f0e7}
body{margin:0;padding:32px}.wrap{max-width:1080px;margin:auto}.flag{display:inline-block;padding:8px 12px;border:2px solid #9d2f2f;background:#fff3f0;color:#8b1f1f;font-weight:800;letter-spacing:.05em}
h1{font-size:32px;margin:18px 0 8px}p{line-height:1.55}.meta{color:#5f675f}.grid{display:grid;grid-template-columns:repeat(auto-fit,minmax(280px,1fr));gap:16px;margin-top:22px}
.card{background:#fff;border:1px solid #d9d2c3;border-radius:14px;padding:18px;box-shadow:0 5px 18px #463a2512}.tag{display:inline-block;border-radius:999px;background:#e7efe8;padding:4px 9px;font-size:12px;font-weight:700;margin-right:6px}
.sponsored{background:#fff0c7;color:#704b00}.unknown{background:#f0e9f8;color:#59356f}a{color:#17633a}dl{display:grid;grid-template-columns:150px 1fr;gap:8px 14px}dt{font-weight:700}dd{margin:0}.warning{padding:14px;border-left:5px solid #d08612;background:#fff8e8}
"""


def page(title: str, body: str) -> str:
    return f"""<!doctype html>
<html lang="en"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>{html.escape(title)}</title><style>{CSS}</style></head>
<body><main class="wrap"><div class="flag">{LABEL}</div>{body}</main></body></html>\n"""


AMAZON_PRODUCTS = [
    {"asin":"TCH-A001","brand":"Grindora","title":"TrailMill Manual Coffee Grinder with Stainless Burr","price":"$59.99","rating":"4.5","reviews":"2,843","bsr":"#18 in Manual Coffee Grinders","note":"BSR visible in teaching detail snapshot"},
    {"asin":"TCH-B002","brand":"RoastPath","title":"Compact Hand Burr Grinder for Travel and Pour Over","price":"$44.50","rating":"4.3","reviews":"918","bsr":"","note":"BSR not visible"},
    {"asin":"TCH-C003","brand":"BeanOrbit","title":"USB Rechargeable Coffee Grinder, 30 Settings","price":"$79.00","rating":"4.1","reviews":"1,204","bsr":"#42 in Electric Burr Coffee Grinders","note":"Different product form; review relevance"},
    {"asin":"TCH-D004","brand":"CampCup","title":"Foldable Handle Manual Grinder, 20g Capacity","price":"$36.99","rating":"4.6","reviews":"5,411","bsr":"#7 in Manual Coffee Grinders","note":"BSR visible in teaching detail snapshot"},
    {"asin":"TCH-E005","brand":"NorthPeak","title":"Aluminum Hand Coffee Mill for Camping","price":"","rating":"4.2","reviews":"377","bsr":"","note":"Price and BSR not visible"},
    {"asin":"TCH-F006","brand":"DailyDose","title":"Ceramic Burr Coffee Grinder with Glass Jar","price":"$25.95","rating":"4.0","reviews":"8,022","bsr":"#31 in Manual Coffee Grinders","note":"BSR visible in teaching detail snapshot"},
    {"asin":"TCH-G007","brand":"Morrow Brew","title":"External-Adjust Manual Grinder for Pour Over","price":"$68.00","rating":"4.7","reviews":"642","bsr":"","note":"BSR not visible"},
    {"asin":"TCH-H008","brand":"NomadCrank","title":"Slim Travel Coffee Grinder with Carry Pouch","price":"$52.49","rating":"","reviews":"","bsr":"","note":"Rating, reviews and BSR not visible"},
    {"asin":"TCH-I009","brand":"BurrWorks","title":"Precision Hand Grinder, 36 Click Adjustment","price":"$89.99","rating":"4.8","reviews":"1,115","bsr":"#12 in Manual Coffee Grinders","note":"BSR visible in teaching detail snapshot"},
    {"asin":"TCH-J010","brand":"HearthRoute","title":"Classic Wooden Coffee Mill with Drawer","price":"$32.00","rating":"4.1","reviews":"3,090","bsr":"","note":"Different use context; BSR not visible"},
    {"asin":"TCH-K011","brand":"PressMate","title":"Hand Grinder for French Press and Cold Brew","price":"$41.75","rating":"4.4","reviews":"706","bsr":"#56 in Manual Coffee Grinders","note":"BSR visible in teaching detail snapshot"},
    {"asin":"TCH-L012","brand":"AeroBean","title":"Lightweight Manual Coffee Grinder, Removable Crank","price":"$47.20","rating":"4.3","reviews":"1,830","bsr":"","note":"BSR not visible"},
]

SEARCH_ITEMS = [
    (1,"Sponsored","TCH-A001"),(2,"Organic","TCH-B002"),(3,"Sponsored","TCH-C003"),(4,"Organic","TCH-D004"),
    (5,"Organic","TCH-B002"),(6,"Organic","TCH-E005"),(7,"Sponsored","TCH-F006"),(8,"Organic","TCH-G007"),
    (9,"Organic","TCH-H008"),(10,"Organic","TCH-I009"),(11,"Organic","TCH-J010"),(12,"Sponsored","TCH-K011"),
    (13,"Organic","TCH-L012"),(14,"Organic","TCH-D004"),(15,"Unknown",""),
]


def generate_amazon() -> None:
    products_dir = AMAZON_ROOT / "products"
    products_dir.mkdir(parents=True, exist_ok=True)
    by_asin = {item["asin"]: item for item in AMAZON_PRODUCTS}
    cards = []
    for position, placement, asin in SEARCH_ITEMS:
        if not asin:
            cards.append(f'<article class="card"><span class="tag unknown">{placement}</span><h2>Position {position}: image-only result card</h2><p>ASIN and product fields are not visible in this teaching card.</p><p class="meta">Source: this local snapshot only</p></article>')
            continue
        item = by_asin[asin]
        tag_class = "sponsored" if placement == "Sponsored" else ""
        cards.append(f'''<article class="card" data-position="{position}" data-placement="{placement}" data-asin="{asin}"><span class="tag {tag_class}">{placement}</span><h2>Position {position}: {html.escape(item["title"])}</h2><p><strong>{html.escape(item["brand"])}</strong> · ASIN {asin}</p><p>Price: {item["price"] or "not visible"} · Rating: {item["rating"] or "not visible"} · Reviews: {item["reviews"] or "not visible"}</p><a href="products/{asin.lower()}.html">Open teaching detail snapshot</a></article>''')
    body = f'''<h1>Teaching search snapshot: portable coffee grinder</h1><p class="meta">Site scenario: Amazon US · Captured for teaching: 2026-08-15 10:30 PDT · 15 visible result cards</p><p class="warning">This layout is fictional. Search position, Sponsored placement and visible BSR are separate signals. The page does not claim sales volume.</p><section class="grid">{"".join(cards)}</section>'''
    (AMAZON_ROOT / "amazon-search-results.html").write_text(page("Amazon teaching search snapshot", body), encoding="utf-8")
    for item in AMAZON_PRODUCTS:
        detail = f'''<h1>{html.escape(item["title"])}</h1><p class="meta">Simplified product detail snapshot · ASIN {item["asin"]}</p><dl><dt>Brand</dt><dd>{html.escape(item["brand"])}</dd><dt>ASIN</dt><dd>{item["asin"]}</dd><dt>Price</dt><dd>{item["price"] or "not visible"}</dd><dt>Rating</dt><dd>{item["rating"] or "not visible"}</dd><dt>Review count</dt><dd>{item["reviews"] or "not visible"}</dd><dt>Visible BSR</dt><dd>{html.escape(item["bsr"] or "not visible")}</dd><dt>Evidence note</dt><dd>{html.escape(item["note"])}</dd></dl><p class="warning">Use only fields visible here. Do not infer missing values or treat BSR as unit sales.</p><p><a href="../amazon-search-results.html">Back to teaching search snapshot</a></p>'''
        (products_dir / f'{item["asin"].lower()}.html').write_text(page(f'{item["asin"]} teaching detail', detail), encoding="utf-8")


INSTAGRAM_PROFILES = [
    ("velvetchair_la","Velvet Chair Wig Salon","Los Angeles, CA","Salon","Custom wig fitting and maintenance. Appointment link in bio.","https://velvetchair.example/book","12.4K","Posts within 7 days"),
    ("laceandlight_studio","Lace & Light Studio","Los Angeles, CA","Salon","Wig installs, styling and consultations.","hello@laceandlight.example","8,210","Posts within 14 days"),
    ("mika_styles_wigs","Mika Styles","Los Angeles, CA","Independent Stylist","Independent wig stylist; booking by public form.","https://mika-styles.example","5,880","Posts within 5 days"),
    ("crownroom_wigs","Crown Room Wigs","Los Angeles, CA","Wig Store","Retail wig showroom and fittings.","https://crownroom.example","21.1K","Posts within 30 days"),
    ("westbeauty_supply","West Beauty Supply","Los Angeles, CA","Beauty Supply","Beauty supply store; wigs are one of many categories.","https://westbeauty.example","3,420","Posts within 10 days"),
    ("aurahair_global","AuraHair Global","","Brand / E-commerce","Ships synthetic wigs worldwide; no salon or local service stated.","https://aurahair.example","98K","Posts within 3 days"),
    ("nia_loves_color","Nia Loves Color","Los Angeles, CA","Consumer / Irrelevant","Personal beauty and lifestyle diary.","","640","Posts within 2 days"),
    ("brooklyn_lace_lab","Brooklyn Lace Lab","Brooklyn, NY","Salon","Wig customization studio in Brooklyn.","https://brooklynlace.example","15.2K","Posts within 8 days"),
    ("styledby_rin","Styled by Rin","","Independent Stylist","Wig styling and installs. Location not stated.","https://rin-booking.example","4,700","Posts within 6 days"),
    ("longbeach_crownbar","Long Beach Crown Bar","Long Beach, CA","Salon","Wig salon and topper consultations.","https://crownbar.example","6,930","Posts within 12 days"),
    ("fadegarage_la","Fade Garage","Los Angeles, CA","Consumer / Irrelevant","Barbershop focused on fades and beard trims; no wig service stated.","https://fadegarage.example","11K","Posts within 4 days"),
    ("private_glam_notes","Private Glam Notes","","Uncertain","Private account; no public business description.","","","Private account"),
    ("thehairarchive","The Hair Archive","","Uncertain","Hair inspiration archive. Business type and location not stated.","","2,080","Last visible post 8 months ago"),
    ("pasadena_wig_atelier","Pasadena Wig Atelier","Pasadena, CA","Salon","Medical and fashion wig consultations by appointment.","https://pasadena-atelier.example","7,315","Posts within 21 days"),
    ("sunset_lace_co","Sunset Lace Co.","Los Angeles, CA","Brand / E-commerce","Online lace wig brand; pickup language is ambiguous.","https://sunsetlace.example","44K","Posts within 4 days"),
    ("beautyclass_mara","Beauty Class with Mara","Los Angeles, CA","Uncertain","Educator teaching wig application; services not stated.","courses@beautyclass.example","9,800","Posts within 9 days"),
    ("wigcart_online","WigCart Online","","Brand / E-commerce","Online-only wig shop; no location or salon service.","https://wigcart.example","31K","Posts within 2 days"),
    ("oldtown_wig_room","Old Town Wig Room","Los Angeles, CA","Salon","Bio says wig salon; public activity appears inactive.","","1,120","Last visible post 19 months ago"),
]


def generate_instagram() -> None:
    profiles_dir = INSTAGRAM_ROOT / "profiles"
    profiles_dir.mkdir(parents=True, exist_ok=True)
    cards = []
    for account, display, location, category, bio, contact, followers, activity in INSTAGRAM_PROFILES:
        cards.append(f'''<article class="card"><span class="tag">Candidate</span><h2>@{account}</h2><p>{html.escape(display)}</p><p>{html.escape(bio)}</p><a href="profiles/{account}.html">Open teaching Profile</a></article>''')
        profile_url = f"https://instagram.example/{account}"
        detail = f'''<h1>@{account}</h1><p class="meta">Fictional public-profile teaching snapshot</p><dl><dt>Display name</dt><dd>{html.escape(display)}</dd><dt>Profile URL</dt><dd>{profile_url}</dd><dt>Location shown</dt><dd>{html.escape(location or "not stated")}</dd><dt>Profile label</dt><dd>{html.escape(category)}</dd><dt>Public bio</dt><dd>{html.escape(bio)}</dd><dt>Website / public contact</dt><dd>{html.escape(contact or "not visible")}</dd><dt>Follower count shown</dt><dd>{html.escape(followers or "not visible")}</dd><dt>Activity signal</dt><dd>{html.escape(activity)}</dd></dl><p class="warning">The label is a teaching hint, not an approved lead decision. Apply the live scope and record why the account is included, rejected or sent to manual review.</p><p><a href="../search-results.html">Back to teaching search results</a></p>'''
        (profiles_dir / f"{account}.html").write_text(page(f"@{account} teaching Profile", detail), encoding="utf-8")
    body = f'''<h1>Teaching search snapshot: wig salon accounts</h1><p class="meta">Location is intentionally mixed or missing · 18 fictional candidates</p><p class="warning">Search results are candidates, not a Lead List. Apply the confirmed geography, target type, exclusions and manual-review rules.</p><section class="grid">{"".join(cards)}</section>'''
    (INSTAGRAM_ROOT / "search-results.html").write_text(page("Instagram teaching search snapshot", body), encoding="utf-8")


def generate_rows() -> list[dict[str, object]]:
    rng = random.Random(20260818)
    start = date(2026, 1, 1)
    markets = ("US", "US", "US", "Canada", "UK")
    channels = ("Amazon Organic", "Amazon Ads", "Shopify", "Wholesale")
    regions = ("West", "Northeast", "South", "Midwest")
    sku_data = {
        "BG-G2-BLK": ("G2 Portable Coffee Grinder - Black", 69.0, 18),
        "BG-G2-SLV": ("G2 Portable Coffee Grinder - Silver", 72.0, 15),
        "BG-G2-CRM": ("G2 Portable Coffee Grinder - Cream", 74.0, 14),
        "BG-M1-BLK": ("M1 Electric Milk Frother - Black", 29.0, 22),
        "BG-T1-GRN": ("T1 Travel Coffee Mug - Green", 26.0, 19),
    }
    rows: list[dict[str, object]] = []
    for i in range(420):
        day_index = i % 90
        current = start + timedelta(days=day_index)
        market = markets[(i * 3 + day_index) % len(markets)]
        channel = channels[(i * 5 + day_index // 3) % len(channels)]
        sku = tuple(sku_data)[(i * 7 + day_index // 5) % len(sku_data)]
        product, price, base = sku_data[sku]
        region = regions[(i * 11 + day_index // 7) % len(regions)]
        channel_factor = {"Amazon Organic":1.15,"Amazon Ads":1.0,"Shopify":0.72,"Wholesale":0.42}[channel]
        market_factor = {"US":1.0,"Canada":0.68,"UK":0.61}[market]
        trend_factor = 1.0
        if sku == "BG-G2-CRM" and current >= date(2026, 2, 15):
            trend_factor = 0.72 if current < date(2026, 3, 1) else 0.46
        orders = max(1, round(base * channel_factor * market_factor * trend_factor + rng.uniform(-3.2, 3.2)))
        units = orders + (1 if rng.random() < 0.24 else 0) + (1 if channel == "Wholesale" and rng.random() < 0.45 else 0)
        revenue = round(units * price * rng.uniform(0.93, 1.01), 2)
        ad_spend = 0.0
        if channel == "Amazon Ads":
            spend_factor = 0.31 if current < date(2026, 3, 1) else 0.57
            ad_spend = round(revenue * spend_factor * rng.uniform(0.94, 1.07), 2)
        return_chance = 0.16 if region == "South" else 0.035
        returns = sum(1 for _ in range(units) if rng.random() < return_chance)
        refund_amount = round(returns * price * rng.uniform(0.78, 0.96), 2)
        sessions: int | None = max(orders, round(orders * rng.uniform(8.0, 15.5)))
        inventory: int | None = max(0, round(510 - day_index * 2.6 + rng.uniform(-28, 32)))
        rows.append({
            "record_id": f"BP-{i+1:04d}", "date": current, "market": market, "channel": channel,
            "sku": sku, "product": product, "orders": orders, "units": units, "revenue": revenue,
            "ad_spend": ad_spend, "refund_amount": refund_amount, "returns": returns,
            "sessions": sessions, "inventory": inventory, "region": region,
        })
    for index in (47, 188, 356, 401):
        rows[index]["sessions"] = None
    for index in (93, 271, 389):
        rows[index]["inventory"] = None
    rows[221]["revenue"] = round(float(rows[221]["revenue"]) * 8.5, 2)
    rows.append(dict(rows[136]))
    return rows


def style_sheet(ws) -> None:
    header_fill = PatternFill("solid", fgColor="1F5B45")
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = Font(color="FFFFFF", bold=True)
        cell.alignment = Alignment(horizontal="center")
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {"A":14,"B":13,"C":12,"D":18,"E":14,"F":40,"G":10,"H":10,"I":14,"J":14,"K":16,"L":10,"M":12,"N":12,"O":13}
    for column, width in widths.items():
        ws.column_dimensions[column].width = width


def generate_workbook(rows: list[dict[str, object]]) -> None:
    DATA_PATH.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    meta = wb.active
    meta.title = "README"
    lines = [
        ("FICTIONAL / TEACHING DATASET", "Not real BrewGo, customer, platform, or market data."),
        ("Purpose", "Day2 Data Analysis → Dashboard classroom lab"),
        ("Data period", "2026-01-01 to 2026-03-31"),
        ("Rows", f"{len(rows)} data rows including one intentional exact duplicate"),
        ("Money", "USD-equivalent teaching values across markets; not accounting statements"),
        ("Known design", "Small number of quality and business signals are intentionally planted; inspect before analysis"),
        ("Input rule", "Do not overwrite this workbook. Write cleaned data and analysis to outputs/analysis.xlsx"),
    ]
    for row in lines:
        meta.append(row)
    meta["A1"].font = Font(size=16, bold=True, color="9C1C1C")
    meta.column_dimensions["A"].width = 24
    meta.column_dimensions["B"].width = 100
    meta.freeze_panes = "A2"

    ws = wb.create_sheet("business_performance")
    headers = list(rows[0])
    ws.append(headers)
    for item in rows:
        ws.append([item[column] for column in headers])
    style_sheet(ws)
    for cell in ws["B"][1:]:
        cell.number_format = "yyyy-mm-dd"
    for column in ("I", "J", "K"):
        for cell in ws[column][1:]:
            cell.number_format = '$#,##0.00'
    table = Table(displayName="BusinessPerformance", ref=ws.dimensions)
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium4", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)
    wb.properties.title = "BrewGo Day2 Business Performance — Fictional Teaching Dataset"
    wb.properties.description = "FICTIONAL / TEACHING DATASET. Not real business data."
    wb.save(DATA_PATH)


def clean_rows(rows: list[dict[str, object]]) -> list[dict[str, object]]:
    seen: set[tuple[object, ...]] = set()
    cleaned = []
    for row in rows:
        key = tuple(row.values())
        if key not in seen:
            cleaned.append(row)
            seen.add(key)
    return cleaned


def aggregate(rows: list[dict[str, object]]) -> dict[str, object]:
    monthly = defaultdict(lambda: {"revenue":0.0,"orders":0,"ad_spend":0.0})
    ads = defaultdict(lambda: {"orders":0,"ad_spend":0.0})
    sku_month = defaultdict(lambda: defaultdict(float))
    region = defaultdict(lambda: {"returns":0,"units":0})
    for row in rows:
        month = row["date"].strftime("%Y-%m")
        monthly[month]["revenue"] += float(row["revenue"])
        monthly[month]["orders"] += int(row["orders"])
        monthly[month]["ad_spend"] += float(row["ad_spend"])
        if row["channel"] == "Amazon Ads":
            ads[month]["orders"] += int(row["orders"])
            ads[month]["ad_spend"] += float(row["ad_spend"])
        sku_month[row["sku"]][month] += float(row["revenue"])
        region[row["region"]]["returns"] += int(row["returns"])
        region[row["region"]]["units"] += int(row["units"])
    return {
        "rows_raw": 421,
        "rows_deduplicated": len(rows),
        "revenue": round(sum(float(row["revenue"]) for row in rows), 2),
        "orders": sum(int(row["orders"]) for row in rows),
        "ad_spend": round(sum(float(row["ad_spend"]) for row in rows), 2),
        "returns": sum(int(row["returns"]) for row in rows),
        "units": sum(int(row["units"]) for row in rows),
        "monthly": {key:{k:round(v,2) if isinstance(v,float) else v for k,v in value.items()} for key,value in sorted(monthly.items())},
        "ads": {key:{k:round(v,2) if isinstance(v,float) else v for k,v in value.items()} for key,value in sorted(ads.items())},
        "sku_month": {sku:{m:round(v,2) for m,v in sorted(values.items())} for sku,values in sorted(sku_month.items())},
        "region": {key:{**value,"return_rate":round(value["returns"] / value["units"],4)} for key,value in sorted(region.items())},
    }


def bars(values: list[float], labels: list[str], color: str, formatter) -> str:
    top = max(values) or 1
    out = []
    for label, value in zip(labels, values):
        width = max(3, round(value / top * 100, 1))
        out.append(f'<div class="bar-row"><span>{html.escape(label)}</span><div class="track"><i style="width:{width}%;background:{color}"></i></div><strong>{formatter(value)}</strong></div>')
    return "".join(out)


def generate_reference_dashboard(metrics: dict[str, object]) -> None:
    months = list(metrics["monthly"])
    revenue_values = [metrics["monthly"][m]["revenue"] for m in months]
    ads_spend = [metrics["ads"][m]["ad_spend"] for m in months]
    ads_orders = [metrics["ads"][m]["orders"] for m in months]
    cream = [metrics["sku_month"]["BG-G2-CRM"][m] for m in months]
    regions = list(metrics["region"])
    return_rates = [metrics["region"][r]["return_rate"] * 100 for r in regions]
    dashboard_css = """
    :root{font-family:Inter,ui-sans-serif,system-ui,-apple-system,"Segoe UI",sans-serif;color:#17231c;background:#f4f1e8}body{margin:0;padding:28px}.wrap{max-width:1280px;margin:auto}h1{font-size:38px;margin:0}.sub{color:#5d665e;margin:8px 0 22px}.flag{display:inline-block;color:#8b1f1f;background:#fff0ed;border:2px solid #a22b2b;padding:7px 11px;font-weight:800}.kpis,.charts{display:grid;gap:16px}.kpis{grid-template-columns:repeat(4,1fr);margin:20px 0}.charts{grid-template-columns:repeat(2,1fr)}.card{background:#fff;border:1px solid #d8d1c2;border-radius:16px;padding:20px;box-shadow:0 7px 22px #4c3d2112}.kpi strong{font-size:29px;display:block}.kpi span{color:#657068}.bar-row{display:grid;grid-template-columns:94px 1fr 92px;align-items:center;gap:10px;margin:13px 0}.track{height:18px;background:#ece9e1;border-radius:999px;overflow:hidden}.track i{display:block;height:100%;border-radius:999px}.bar-row strong{text-align:right;font-size:13px}.note{border-left:5px solid #d28513;background:#fff8e7;padding:14px;margin-top:18px}.trace{width:100%;border-collapse:collapse;font-size:14px}.trace th,.trace td{padding:9px;border-bottom:1px solid #e2ddd2;text-align:left}.trace th{background:#eef3ed}@media(max-width:850px){.kpis,.charts{grid-template-columns:1fr 1fr}}@media(max-width:560px){.kpis,.charts{grid-template-columns:1fr}.bar-row{grid-template-columns:76px 1fr 78px}}
    """
    body = f'''<div class="wrap"><div class="flag">FICTIONAL / TEACHING REFERENCE</div><h1>BrewGo business performance</h1><p class="sub">Instructor Plan B · 2026-01-01 to 2026-03-31 · exact duplicate removed; revenue outlier retained and disclosed</p>
    <section class="kpis"><div class="card kpi"><strong>${metrics["revenue"]:,.0f}</strong><span>Revenue · USD-equivalent</span></div><div class="card kpi"><strong>{metrics["orders"]:,}</strong><span>Orders</span></div><div class="card kpi"><strong>${metrics["ad_spend"]:,.0f}</strong><span>Ad spend</span></div><div class="card kpi"><strong>{metrics["returns"]/metrics["units"]:.1%}</strong><span>Returns / units</span></div></section>
    <section class="charts"><article class="card"><h2>收入是否在三个月内持续增长？</h2>{bars(revenue_values,months,"#1f6b4a",lambda x:f"${x/1000:.1f}K")}</article><article class="card"><h2>Amazon Ads 花费上升时，订单同步了吗？</h2>{bars(ads_spend,months,"#cb7c17",lambda x:f"${x:,.0f}")}<hr>{bars(ads_orders,months,"#466a9f",lambda x:f"{int(x):,} 单")}</article><article class="card"><h2>BG-G2-CRM 的收入是否在走弱？</h2>{bars(cream,months,"#8b5ca6",lambda x:f"${x/1000:.1f}K")}</article><article class="card"><h2>哪个区域的单位退货率最高？</h2>{bars(return_rates,regions,"#b34b3f",lambda x:f"{x:.1f}%")}</article></section>
    <p class="note"><strong>质量说明：</strong>原始 421 行含 1 条精确重复、4 个 sessions 空值、3 个 inventory 空值和 1 个收入异常高值。参考口径仅去除精确重复；异常高值保留，因此收入趋势需要敏感性复核。同步变化不能证明因果。</p>
    <article class="card"><h2>关键数字追溯</h2><table class="trace"><thead><tr><th>指标</th><th>口径</th><th>来源</th></tr></thead><tbody><tr><td>Revenue / Orders / Ad spend</td><td>精确去重后求和；异常收入保留</td><td>business_performance sheet</td></tr><tr><td>单位退货率</td><td>SUM(returns) / SUM(units)</td><td>按 region 汇总</td></tr><tr><td>广告图</td><td>channel = Amazon Ads，按月汇总</td><td>date、orders、ad_spend</td></tr><tr><td>CRM 趋势</td><td>sku = BG-G2-CRM，按月收入</td><td>date、sku、revenue</td></tr></tbody></table></article></div>'''
    REFERENCE_DASHBOARD.write_text(f'<!doctype html><html lang="zh-CN"><head><meta charset="utf-8"><meta name="viewport" content="width=device-width,initial-scale=1"><title>Day2 Data Dashboard Reference</title><style>{dashboard_css}</style></head><body>{body}</body></html>\n', encoding="utf-8")
    REFERENCE_METRICS.write_text(json.dumps(metrics, ensure_ascii=False, indent=2, sort_keys=True) + "\n", encoding="utf-8")


def verify_workbook() -> None:
    wb = load_workbook(DATA_PATH, read_only=True, data_only=True)
    if wb.sheetnames != ["README", "business_performance"]:
        raise SystemExit(f"Unexpected workbook sheets: {wb.sheetnames}")
    ws = wb["business_performance"]
    if ws.max_row != 422 or ws.max_column != 15:
        raise SystemExit(f"Unexpected workbook shape: {ws.max_row}x{ws.max_column}")
    if "FICTIONAL / TEACHING DATASET" not in str(wb["README"]["A1"].value):
        raise SystemExit("Workbook teaching label missing")


def main() -> None:
    if DAY2_ROOT.resolve().parent != (ROOT / "classroom").resolve():
        raise SystemExit(f"Unsafe Day2 fixture root: {DAY2_ROOT}")
    generate_amazon()
    generate_instagram()
    rows = generate_rows()
    generate_workbook(rows)
    verify_workbook()
    metrics = aggregate(clean_rows(rows))
    generate_reference_dashboard(metrics)
    print(f"Generated Amazon snapshots: search + {len(AMAZON_PRODUCTS)} product details")
    print(f"Generated Instagram snapshots: search + {len(INSTAGRAM_PROFILES)} profiles")
    print(f"Generated data workbook: {len(rows)} rows, 15 columns")
    print(f"Generated offline instructor dashboard: {REFERENCE_DASHBOARD.relative_to(ROOT)}")


if __name__ == "__main__":
    main()
