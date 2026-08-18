#!/usr/bin/env python3
"""Validate BrewGo Day2 Live Tasks classroom assets and generated workspaces."""
from __future__ import annotations

import hashlib
import json
import re
import statistics
from collections import Counter, defaultdict
from html.parser import HTMLParser
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
CLASSROOM = ROOT / "classroom" / "day2-live-tasks"
WORKSPACES = ROOT / "workspaces" / "codex" / "day2-live-tasks"
INSTRUCTOR = ROOT / "instructor"
PROTECTION_RECEIPT = WORKSPACES / "protected-assets-manifest.json"
TASKS = (
    "01-amazon-competitor-discovery",
    "02-instagram-lead-discovery",
    "03-data-analysis-dashboard",
)
errors: list[str] = []
passes: list[str] = []


class AuditHTMLParser(HTMLParser):
    def __init__(self) -> None:
        super().__init__()
        self.tags: Counter[str] = Counter()
        self.attrs: list[tuple[str, dict[str, str]]] = []
        self.text: list[str] = []

    def handle_starttag(self, tag: str, attrs: list[tuple[str, str | None]]) -> None:
        values = {key: value or "" for key, value in attrs}
        self.tags[tag] += 1
        self.attrs.append((tag, values))

    def handle_data(self, data: str) -> None:
        self.text.append(data)


def check(condition: bool, message: str) -> None:
    if condition:
        passes.append(message)
    else:
        errors.append(message)


def parse_html(path: Path) -> AuditHTMLParser:
    parser = AuditHTMLParser()
    try:
        parser.feed(path.read_text(encoding="utf-8"))
        parser.close()
    except Exception as exc:  # noqa: BLE001 - validation should aggregate failures
        errors.append(f"HTML parse failed: {path}: {exc}")
    return parser


def external_dependencies(parser: AuditHTMLParser) -> list[str]:
    external: list[str] = []
    for tag, attrs in parser.attrs:
        if tag not in {"script", "link", "img", "iframe"}:
            continue
        value = attrs.get("src") or attrs.get("href") or ""
        if urlparse(value).scheme in {"http", "https"}:
            external.append(value)
    return external


def validate_local_links(path: Path, parser: AuditHTMLParser) -> None:
    for tag, attrs in parser.attrs:
        if tag != "a":
            continue
        href = attrs.get("href", "")
        parsed = urlparse(href)
        if not href or href.startswith("#") or parsed.scheme in {"http", "https", "mailto"}:
            continue
        target = (path.parent / parsed.path).resolve()
        check(target.is_file(), f"local link resolves: {path.name} -> {href}")


def amazon_search_stats(path: Path) -> dict[str, object]:
    parser = parse_html(path)
    validate_local_links(path, parser)
    raw = path.read_text(encoding="utf-8")
    cards = [
        attrs
        for tag, attrs in parser.attrs
        if tag == "article" and "card" in attrs.get("class", "").split()
    ]
    placements = Counter(card.get("data-placement", "") for card in cards)
    asins = [card.get("data-asin", "") for card in cards if card.get("data-asin")]
    organic_asins = [
        card.get("data-asin", "")
        for card in cards
        if card.get("data-placement") == "Organic" and card.get("data-asin")
    ]
    return {
        "parser": parser,
        "text": " ".join(parser.text),
        "total": len(cards),
        "sponsored": placements["Sponsored"],
        "organic": placements["Organic"],
        "unique_organic": len(set(organic_asins)),
        "duplicate": sum(count - 1 for count in Counter(asins).values() if count > 1),
        "unknown": placements["Unknown"],
        "missing": sum(card.get("data-missing") == "true" for card in cards),
        "asins": set(asins),
        "brands": {card.get("data-brand", "") for card in cards if card.get("data-brand")},
        "prices": {card.get("data-price", "") for card in cards if card.get("data-price")},
        "titles": {
            value.strip()
            for value in re.findall(r"<h2>Position \d+: ([^<]+)</h2>", raw, flags=re.IGNORECASE)
            if "image-only result card" not in value.lower()
        },
    }


def snapshot_paths(paths: tuple[Path, ...]) -> dict[str, str]:
    protected: dict[str, str] = {}
    for root in paths:
        candidates = [root] if root.is_file() else sorted(root.rglob("*")) if root.is_dir() else []
        for path in candidates:
            if path.is_file():
                digest = hashlib.sha256(path.read_bytes()).hexdigest()
                protected[path.relative_to(ROOT).as_posix()] = digest
    return protected


def demo_snapshot() -> dict[str, str]:
    paths: list[Path] = []
    for base, pattern in ((ROOT / "classroom", "0[1-7]-*"), (ROOT / "workspaces" / "codex", "0[1-7]-*"), (INSTRUCTOR, "demo-0*")):
        paths.extend(sorted(base.glob(pattern)))
    return snapshot_paths(tuple(paths))


def protected_asset_snapshots() -> dict[str, dict[str, str]]:
    return {
        "Demo01-07": demo_snapshot(),
        "Quick Wins": snapshot_paths((
            ROOT / "classroom" / "quick-wins",
            ROOT / "workspaces" / "codex" / "quick-wins",
            INSTRUCTOR / "quick-wins-control.html",
            INSTRUCTOR / "quick-wins-runbook.md",
            ROOT / "scripts" / "build_quick_wins.py",
        )),
        "Data Analysis": snapshot_paths((
            CLASSROOM / TASKS[2],
            WORKSPACES / TASKS[2],
            INSTRUCTOR / "day2-data-dashboard-reference.html",
            INSTRUCTOR / "day2-data-dashboard-reference-metrics.json",
        )),
    }


def snapshot_digest(files: dict[str, str]) -> str:
    payload = json.dumps(files, ensure_ascii=False, sort_keys=True, separators=(",", ":"))
    return hashlib.sha256(payload.encode("utf-8")).hexdigest()


def validate_source_and_workspaces() -> None:
    check(CLASSROOM.is_dir(), "classroom/day2-live-tasks exists")
    check(WORKSPACES.is_dir(), "generated Day2 workspace root exists")
    check(tuple(sorted(path.name for path in WORKSPACES.iterdir() if path.is_dir())) == TASKS, "three generated workspaces match whitelist")
    rules = (CLASSROOM / "WORKSPACE_RULES.md").read_text(encoding="utf-8")
    check("不得覆盖" in rules and "只写入 `outputs/`" in rules, "common rules protect original inputs and constrain outputs")
    check("Live 数据与 Offline 教学数据必须明确区分" in rules, "common rules separate Live and Offline modes")

    for task in TASKS:
        source = CLASSROOM / task
        workspace = WORKSPACES / task
        for name in ("README.md", "RAW_REQUEST.md", "AGENTS.md", "input", "outputs", "workspace-manifest.json"):
            check((workspace / name).exists(), f"{task} workspace has {name}")
        check(not (workspace / "INSTRUCTOR_GUIDE.md").exists(), f"{task} excludes instructor guide")
        check(not list(workspace.rglob("SKILL.md")) and not list(workspace.rglob(".agents")), f"{task} has no preinstalled final Skill")
        output_entries = [path.name for path in (workspace / "outputs").iterdir() if path.name != ".gitkeep"]
        check(not output_entries, f"{task} workspace outputs initially empty")
        source_output_entries = [path.name for path in (source / "outputs").iterdir() if path.name != ".gitkeep"]
        check(not source_output_entries, f"{task} source outputs initially empty")
        manifest = json.loads((workspace / "workspace-manifest.json").read_text(encoding="utf-8"))
        check(manifest.get("outputs_initially_empty") is True and manifest.get("instructor_assets_excluded") is True, f"{task} manifest records isolation contract")
        for source_input in sorted((source / "input").rglob("*")):
            if source_input.is_file():
                relative = source_input.relative_to(source / "input")
                target = workspace / "input" / relative
                check(target.is_file() and target.read_bytes() == source_input.read_bytes(), f"{task} input copied without mutation: {relative}")


def validate_offline_snapshots() -> None:
    amazon_input = WORKSPACES / TASKS[0] / "input"
    amazon_first_root = amazon_input / "offline"
    amazon_second_root = amazon_input / "offline-second-run"
    first_html = sorted(amazon_first_root.rglob("*.html"))
    second_html = sorted(amazon_second_root.rglob("*.html"))
    check(amazon_first_root.is_dir(), "Amazon first-run Offline exists")
    check(amazon_second_root.is_dir(), "Amazon second-run Offline exists")
    check(len(first_html) == 16, "Amazon first-run Offline contains one search page and 15 detail pages")
    check(len(second_html) == 14, "Amazon second-run Offline contains one search page and 13 detail pages")

    first = amazon_search_stats(amazon_first_root / "amazon-search-results.html")
    second = amazon_search_stats(amazon_second_root / "amazon-search-results.html")
    check("portable coffee grinder" in first["text"].lower(), "Amazon first-run keyword is portable coffee grinder")
    check(first["sponsored"] >= 3, "Amazon first-run has at least three Sponsored cards")
    check(first["organic"] >= 12, "Amazon first-run has at least twelve Organic card entries")
    check(first["unique_organic"] >= 10, "Amazon first-run has at least ten unique Organic products after ASIN deduplication")
    check(first["duplicate"] >= 2, "Amazon first-run contains at least two duplicate product entries")
    check(first["unknown"] >= 1 and first["missing"] >= 1, "Amazon first-run contains unknown and missing-data cards")
    check(first["unique_organic"] >= 10, "Amazon first-run can form Organic Top10 after excluding Sponsored and deduplicating ASIN")

    check("electric milk frother" in second["text"].lower(), "Amazon second-run keyword is electric milk frother")
    check("portable coffee grinder" not in second["text"].lower(), "Amazon second-run keyword differs from first-run")
    check(second["sponsored"] >= 3, "Amazon second-run contains Sponsored cards")
    check(second["unique_organic"] >= 10, "Amazon second-run has at least ten unique Organic products")
    check(second["duplicate"] >= 1, "Amazon second-run contains a duplicate product entry")
    check(second["missing"] >= 1, "Amazon second-run contains missing-data cards")
    check(not first["asins"].intersection(second["asins"]), "Amazon first-run and second-run product IDs do not overlap")
    check(not first["brands"].intersection(second["brands"]), "Amazon first-run and second-run brands do not overlap")
    check(not first["prices"].intersection(second["prices"]), "Amazon first-run and second-run visible prices do not overlap")

    bsr_categories: set[str] = set()
    for path in (*first_html, *second_html):
        parser = parse_html(path)
        validate_local_links(path, parser)
        page_text = " ".join(parser.text)
        check("FICTIONAL / TEACHING SNAPSHOT" in page_text, f"Amazon snapshot labeled: {path.parent.parent.name}/{path.name}")
        check(not external_dependencies(parser), f"Amazon snapshot has no external runtime dependency: {path.name}")
        raw = path.read_text(encoding="utf-8")
        bsr_categories.update(re.findall(r"<dd>#\d+\s+in\s+([^<]+)</dd>", raw, flags=re.IGNORECASE))
    check("Manual Coffee Grinders" in bsr_categories and "Electric Burr Coffee Grinders" in bsr_categories, "Amazon Offline preserves at least two distinct BSR categories")

    instagram_root = WORKSPACES / TASKS[1] / "input" / "offline"
    instagram_html = sorted(instagram_root.rglob("*.html"))
    check(len(instagram_html) == 19, "Instagram Offline contains one search page and 18 Profile pages")
    instagram_search = parse_html(instagram_root / "search-results.html")
    validate_local_links(instagram_root / "search-results.html", instagram_search)
    instagram_text = " ".join(instagram_search.text)
    profile_cards = sum(1 for tag, attrs in instagram_search.attrs if tag == "article" and "card" in attrs.get("class", "").split())
    check(profile_cards == 18, "Instagram teaching search page has 18 candidate cards")
    check("FICTIONAL / TEACHING SNAPSHOT" in instagram_text, "Instagram search page visibly labels teaching snapshot")
    for path in instagram_html:
        parser = parse_html(path)
        validate_local_links(path, parser)
        check("FICTIONAL / TEACHING SNAPSHOT" in " ".join(parser.text), f"Instagram snapshot labeled: {path.name}")
        check(not external_dependencies(parser), f"Instagram snapshot has no external runtime dependency: {path.name}")

    profile_paths = sorted((instagram_root / "profiles").glob("*.html"))
    forbidden_labels = {"profile label", "category", "expected category", "account type", "classification", "decision", "manual review expected"}
    for path in profile_paths:
        raw = path.read_text(encoding="utf-8")
        dt_labels = {re.sub(r"<[^>]+>", "", value).strip().lower() for value in re.findall(r"<dt>(.*?)</dt>", raw, flags=re.IGNORECASE | re.DOTALL)}
        parser = parse_html(path)
        structured_attrs = {
            key.lower()
            for _tag, attrs in parser.attrs
            for key in attrs
            if key.lower() in {"data-category", "data-classification", "data-decision", "data-expected-category"}
        }
        check("profile label" not in raw.lower(), f"Instagram student Profile omits Profile label: {path.name}")
        check(not dt_labels.intersection(forbidden_labels) and not structured_attrs, f"Instagram student Profile omits structured classification answers: {path.name}")

    reference_path = INSTRUCTOR / "instagram-offline-reference.json"
    check(reference_path.is_file(), "Instagram instructor classification reference exists")
    if reference_path.is_file():
        reference = json.loads(reference_path.read_text(encoding="utf-8"))
        profiles = reference.get("profiles", [])
        required = {"account", "expected_category", "expected_decision", "reason", "manual_review_expected"}
        check(len(profiles) == 18 and all(required.issubset(item) for item in profiles), "Instagram instructor reference retains 18 complete classification answers")
    check(not (WORKSPACES / TASKS[1] / "instagram-offline-reference.json").exists() and not list((WORKSPACES / TASKS[1]).rglob("instagram-offline-reference.json")), "Instagram instructor classification reference is excluded from student workspace")

    instagram_readme = (WORKSPACES / TASKS[1] / "README.md").read_text(encoding="utf-8")
    check(all(phrase in instagram_readme for phrase in ("不自动关注", "不自动私信", "不做批量营销", "不访问私人内容")), "Instagram task explicitly forbids follow, DM, bulk marketing, and private access")


def read_data_rows() -> list[dict[str, object]]:
    workbook = WORKSPACES / TASKS[2] / "input" / "business-performance.xlsx"
    wb = load_workbook(workbook, read_only=True, data_only=True)
    check(wb.sheetnames == ["README", "business_performance"], "data workbook has README and business_performance sheets")
    check("FICTIONAL / TEACHING DATASET" in str(wb["README"]["A1"].value), "data workbook visibly labels teaching dataset")
    ws = wb["business_performance"]
    values = ws.iter_rows(values_only=True)
    headers = next(values)
    rows = [dict(zip(headers, row)) for row in values]
    check(len(rows) == 421 and len(headers) == 15, "data workbook has 421 rows and 15 fields")
    return rows


def validate_dataset() -> None:
    rows = read_data_rows()
    keys = [tuple(row.values()) for row in rows]
    duplicates = sum(count - 1 for count in Counter(keys).values() if count > 1)
    check(duplicates == 1, "dataset contains exactly one exact duplicate")
    check(sum(row["sessions"] is None for row in rows) == 4, "dataset contains four missing sessions values")
    check(sum(row["inventory"] is None for row in rows) == 3, "dataset contains three missing inventory values")
    revenue = [float(row["revenue"]) for row in rows]
    check(max(revenue) > statistics.median(revenue) * 5, "dataset contains a detectable high revenue outlier")

    unique: list[dict[str, object]] = []
    seen: set[tuple[object, ...]] = set()
    for row in rows:
        key = tuple(row.values())
        if key not in seen:
            unique.append(row)
            seen.add(key)
    ads = defaultdict(lambda: {"spend": 0.0, "orders": 0})
    cream = defaultdict(float)
    regions = defaultdict(lambda: {"returns": 0, "units": 0})
    for row in unique:
        month = row["date"].strftime("%Y-%m")
        if row["channel"] == "Amazon Ads":
            ads[month]["spend"] += float(row["ad_spend"])
            ads[month]["orders"] += int(row["orders"])
        if row["sku"] == "BG-G2-CRM":
            cream[month] += float(row["revenue"])
        regions[row["region"]]["returns"] += int(row["returns"])
        regions[row["region"]]["units"] += int(row["units"])
    check(ads["2026-03"]["spend"] > ads["2026-02"]["spend"] and ads["2026-03"]["orders"] <= ads["2026-02"]["orders"], "Amazon Ads spend rises from February to March without order growth")
    check(cream["2026-03"] < cream["2026-01"] * 0.6, "BG-G2-CRM shows a material sales decline by March")
    return_rates = {region: values["returns"] / values["units"] for region, values in regions.items()}
    check(max(return_rates, key=return_rates.get) == "South" and return_rates["South"] > 0.1, "South has the highest and materially elevated return rate")


def validate_instructor_assets() -> None:
    runbook = INSTRUCTOR / "day2-live-tasks-runbook.md"
    acceptance = INSTRUCTOR / "day2-live-tasks-acceptance.md"
    skill = INSTRUCTOR / "amazon-skill-reference.md"
    control = INSTRUCTOR / "day2-live-tasks-control.html"
    dashboard = INSTRUCTOR / "day2-data-dashboard-reference.html"
    for path in (runbook, acceptance, skill, control, dashboard):
        check(path.is_file(), f"instructor asset exists: {path.name}")
    runbook_text = runbook.read_text(encoding="utf-8")
    check(runbook_text.count("【原始需求】") == 3 and runbook_text.count("【一句收口】") == 3, "runbook uses required compact sections for all three tasks")
    check(all(title in runbook_text for title in ("【如何进入 Skill Lab】", "【第二关键词怎么选】", "【Skill 验收重点】")), "Amazon runbook contains Skill Lab sections")
    skill_text = skill.read_text(encoding="utf-8")
    required_skill_sections = ("## Trigger", "## Inputs", "## Scope", "## Discovery workflow", "## Sponsored handling", "## Deduplication", "## Field extraction", "## Evidence rules", "## Missing-data rules", "## Output contract", "## Human review", "## Self-check")
    check(all(section in skill_text for section in required_skill_sections), "Amazon instructor Skill Plan B covers complete method contract")
    first = amazon_search_stats(WORKSPACES / TASKS[0] / "input" / "offline" / "amazon-search-results.html")
    second = amazon_search_stats(WORKSPACES / TASKS[0] / "input" / "offline-second-run" / "amazon-search-results.html")
    product_results = first["asins"] | second["asins"] | first["brands"] | second["brands"] | first["prices"] | second["prices"] | first["titles"] | second["titles"]
    check(not any(value and value in skill_text for value in product_results), "Amazon Skill Plan B does not hardcode first-run or second-run product results")

    amazon_readme = (CLASSROOM / TASKS[0] / "README.md").read_text(encoding="utf-8")
    amazon_guide = (CLASSROOM / TASKS[0] / "INSTRUCTOR_GUIDE.md").read_text(encoding="utf-8")
    acceptance_text = acceptance.read_text(encoding="utf-8")
    bsr_docs = {
        "Amazon README": amazon_readme,
        "Amazon instructor guide": amazon_guide,
        "Amazon Skill Plan B": skill_text,
        "Day2 runbook": runbook_text,
        "Day2 acceptance": acceptance_text,
    }
    for name, content in bsr_docs.items():
        check("BSR" in content and "具体类目" in content and "不得直接横向比较" in content and "销量件数" in content, f"{name} states the cross-category BSR boundary")

    skill_path = ".agents/skills/amazon-competitor-discovery/SKILL.md"
    skill_path_docs = {
        "Amazon README": amazon_readme,
        "Amazon instructor guide": amazon_guide,
        "Amazon Skill Plan B": skill_text,
        "Day2 runbook": runbook_text,
        "Day2 acceptance": acceptance_text,
    }
    for name, content in skill_path_docs.items():
        check(skill_path in content, f"{name} states the formal Skill path")
    check("input/offline-second-run/" in amazon_readme and "electric milk frother" in amazon_readme, "Amazon README documents second-run Offline fallback")
    check("input/offline-second-run/" in amazon_guide and "electric milk frother" in amazon_guide, "Amazon instructor guide documents second-run Offline fallback")
    check("input/offline-second-run/" in runbook_text and "electric milk frother" in runbook_text, "Day2 runbook documents second-run Offline fallback")

    control_parser = parse_html(control)
    validate_local_links(control, control_parser)
    control_text = " ".join(control_parser.text)
    tab_buttons = sum(1 for tag, attrs in control_parser.attrs if tag == "button" and "tab" in attrs.get("class", "").split())
    check(tab_buttons == 3, "control page has three independent task tabs")
    check(all(label in control_text for label in ("原始需求", "澄清", "第一次执行", "Skill", "新关键词复跑", "验收")), "control page shows complete Amazon learning flow")
    check(skill_path in control_text and "input/offline-second-run/amazon-search-results.html" in control_text, "control page shows formal Skill path and second-run Offline path")
    check(not external_dependencies(control_parser), "control page has no external runtime dependency")

    dashboard_parser = parse_html(dashboard)
    dashboard_text = " ".join(dashboard_parser.text)
    chart_articles = sum(1 for tag, attrs in dashboard_parser.attrs if tag == "article" and "card" in attrs.get("class", "").split()) - 1
    check(chart_articles >= 4, "reference Dashboard contains at least four business charts")
    check("关键数字追溯" in dashboard_text and "同步变化不能证明因果" in dashboard_text, "reference Dashboard includes traceability and causality boundary")
    check(not external_dependencies(dashboard_parser), "reference Dashboard is single-file with no online runtime dependency")


def validate_protected_assets() -> None:
    check(PROTECTION_RECEIPT.is_file(), "protected asset hash receipt exists")
    if not PROTECTION_RECEIPT.is_file():
        return
    receipt = json.loads(PROTECTION_RECEIPT.read_text(encoding="utf-8"))
    current = protected_asset_snapshots()
    for name in ("Demo01-07", "Quick Wins", "Data Analysis"):
        expected = receipt.get(name, {})
        check(expected.get("file_count") == len(current[name]), f"{name} protected file count unchanged")
        check(expected.get("sha256") == snapshot_digest(current[name]), f"{name} protected asset hash unchanged")


def main() -> None:
    validate_source_and_workspaces()
    validate_offline_snapshots()
    validate_dataset()
    validate_instructor_assets()
    validate_protected_assets()
    print(f"PASS {len(passes)} checks")
    for message in passes:
        print(f"  PASS {message}")
    if errors:
        print(f"FAIL {len(errors)} checks")
        for message in errors:
            print(f"  FAIL {message}")
        raise SystemExit(1)
    print("ALL DAY2 LIVE TASK CHECKS PASSED")


if __name__ == "__main__":
    main()
