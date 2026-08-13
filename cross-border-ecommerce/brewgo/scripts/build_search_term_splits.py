#!/usr/bin/env python3
"""Deterministically split data/raw/search_terms.xlsx into two classroom sets.

Derives two teaching inputs for Demo 05 (Amazon Ads Search Term Analysis Skill):

- classroom/05-search-term-skill/input/history/search_terms_history.xlsx
  The earlier observation periods, used for the first full hand-off.
- classroom/05-search-term-skill/input/next-period/search_terms_latest.xlsx
  The later observation periods (rolling window), used for the Skill re-run.

The split is a rolling window over the five observation periods present in the
raw file. The raw date_range strings are intentionally inconsistent (ISO, English
month, slash) and are preserved verbatim as a registered teaching issue.

This script only reads data/raw; it never modifies raw or expected. It requires
openpyxl. Re-run it after any change to data/raw/search_terms.xlsx to keep the
committed classroom inputs in sync.
"""
from __future__ import annotations

import zipfile
from pathlib import Path

from openpyxl import Workbook, load_workbook

SOURCE_ROOT = Path(__file__).resolve().parents[1]
RAW = SOURCE_ROOT / "data" / "raw" / "search_terms.xlsx"
OUT_ROOT = SOURCE_ROOT / "classroom" / "05-search-term-skill" / "input"

# Fixed DOS timestamp used when re-writing the output xlsx zip entries, so the
# derived files are byte-identical across runs (openpyxl stamps the current time
# otherwise, which only churns zip metadata).
_FIXED_ZIP_TS = (1980, 1, 1, 0, 0, 0)

# The five observation periods, in chronological order, keyed by their exact
# date_range string (including the en-dash in the English-month period).
_P1 = "2026-06-01 to 2026-06-14"
_P2 = "Jun 15\u2013Jun 28, 2026"
_P3 = "2026/06/29-2026/07/12"
_P4 = "2026-07-13 to 2026-07-26"
_P5 = "2026-07-27 to 2026-08-08"

HISTORY_PERIODS = (_P1, _P2, _P3)
NEXT_PERIODS = (_P4, _P5)

OUTPUTS = (
    (Path("history/search_terms_history.xlsx"), HISTORY_PERIODS),
    (Path("next-period/search_terms_latest.xlsx"), NEXT_PERIODS),
)


def normalize_zip_timestamps(path: Path) -> None:
    """Rewrite an xlsx with fixed zip entry timestamps for byte-reproducibility."""
    tmp = path.with_suffix(".tmp.xlsx")
    with zipfile.ZipFile(path, "r") as src, zipfile.ZipFile(tmp, "w", zipfile.ZIP_DEFLATED) as dst:
        for info in src.infolist():
            entry = zipfile.ZipInfo(info.filename, _FIXED_ZIP_TS)
            entry.compress_type = info.compress_type
            entry.external_attr = info.external_attr
            entry.create_system = info.create_system
            dst.writestr(entry, src.read(info.filename))
    tmp.replace(path)


def main() -> None:
    wb = load_workbook(RAW, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    header, data = rows[0], rows[1:]

    all_periods = set(HISTORY_PERIODS) | set(NEXT_PERIODS)
    present = {row[0] for row in data}
    if not all_periods <= present:
        missing = all_periods - present
        raise SystemExit(f"raw file is missing expected periods: {sorted(missing)}")

    split: dict[str, list[tuple]] = {}
    for rel, periods in OUTPUTS:
        bucket = [row for row in data if row[0] in periods]
        split[str(rel)] = bucket

    # Partition checks: every raw row lands in exactly one bucket, no overlap.
    history_ids = {id(row) for row in split["history/search_terms_history.xlsx"]}
    next_ids = {id(row) for row in split["next-period/search_terms_latest.xlsx"]}
    if history_ids & next_ids:
        raise SystemExit("split produced overlapping rows")
    if len(history_ids) + len(next_ids) != len(data):
        raise SystemExit("split did not cover all raw rows")

    OUT_ROOT.mkdir(parents=True, exist_ok=True)
    for rel, periods in OUTPUTS:
        target = OUT_ROOT / rel
        target.parent.mkdir(parents=True, exist_ok=True)
        out = Workbook()
        sheet = out.active
        sheet.title = "Search Terms"
        sheet.append(list(header))
        for row in split[str(rel)]:
            sheet.append(["" if value is None else value for value in row])
        out.save(target)
        normalize_zip_timestamps(target)

    print("Search Term classroom split written:")
    for rel, periods in OUTPUTS:
        bucket = split[str(rel)]
        period_counts = {p: sum(1 for r in bucket if r[0] == p) for p in periods}
        print(f"  {rel}: {len(bucket)} rows")
        for period in periods:
            print(f"    - {period}: {period_counts[period]} rows")


if __name__ == "__main__":
    main()
