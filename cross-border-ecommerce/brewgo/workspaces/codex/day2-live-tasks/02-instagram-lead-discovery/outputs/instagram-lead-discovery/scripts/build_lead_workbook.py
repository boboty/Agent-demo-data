#!/usr/bin/env python3
"""Create or append a Chinese Instagram lead workbook from validated JSON."""

from __future__ import annotations

import argparse
import json
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo


HEADERS = [
    "账号", "显示名称", "Instagram 主页", "所在地", "账号类型", "简介摘要",
    "网站 / 公开联系方式", "粉丝数", "内容可见性信号", "入选理由",
    "风险 / 排除理由", "来源", "置信度", "人工复核",
]
KEYS = [
    "account", "display_name", "profile_url", "location", "category", "bio_summary",
    "public_contact", "follower_count", "activity_signal", "why_matched", "risk",
    "source", "confidence", "manual_review",
]
WIDTHS = [25, 30, 44, 34, 29, 48, 55, 19, 44, 54, 48, 60, 14, 46]
HEADER_ROW = 4
DATA_ROW = 5


def canonical_handle(value: str) -> str:
    return value.strip().lstrip("@").casefold()


def normalize_source(value: Any) -> str:
    if isinstance(value, list):
        return "\n".join(str(item).strip() for item in value if str(item).strip())
    return str(value or "").strip()


def load_leads(path: Path) -> list[dict[str, Any]]:
    data = json.loads(path.read_text(encoding="utf-8"))
    if not isinstance(data, list) or not data:
        raise ValueError("Input must be a non-empty JSON array.")
    seen: set[str] = set()
    result: list[dict[str, Any]] = []
    for index, item in enumerate(data, 1):
        if not isinstance(item, dict):
            raise ValueError(f"Lead {index} must be an object.")
        missing = [key for key in KEYS if key not in item]
        if missing:
            raise ValueError(f"Lead {index} is missing keys: {', '.join(missing)}")
        handle = canonical_handle(str(item["account"]))
        if not handle:
            raise ValueError(f"Lead {index} has an empty account.")
        if handle in seen:
            raise ValueError(f"Duplicate input account: @{handle}")
        if str(item["confidence"]) not in {"高", "中", "低"}:
            raise ValueError(f"Lead {index} confidence must be 高, 中, or 低.")
        seen.add(handle)
        row = dict(item)
        row["account"] = "@" + handle
        row["source"] = normalize_source(row["source"])
        result.append(row)
    return result


def style_title(ws, title: str, subtitle: str, color: str = "243B53") -> None:
    ws.sheet_view.showGridLines = False
    ws.freeze_panes = "D5"
    ws.merge_cells("A1:N1")
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=16, color="FFFFFF")
    ws["A1"].fill = PatternFill("solid", fgColor=color)
    ws["A1"].alignment = Alignment(vertical="center")
    ws.row_dimensions[1].height = 28
    ws.merge_cells("A2:N2")
    ws["A2"] = subtitle
    ws["A2"].font = Font(italic=True, color=color)
    ws["A2"].fill = PatternFill("solid", fgColor="EAF2F8")
    ws["A2"].alignment = Alignment(wrap_text=True, vertical="center")
    ws.row_dimensions[2].height = 32


def style_headers(ws, color: str = "486581") -> None:
    for column, header in enumerate(HEADERS, 1):
        cell = ws.cell(HEADER_ROW, column, header)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor=color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.column_dimensions[get_column_letter(column)].width = WIDTHS[column - 1]
    ws.row_dimensions[HEADER_ROW].height = 32


def write_rows(ws, leads: list[dict[str, Any]], start_row: int) -> None:
    thin = Side(style="thin", color="D9E2EC")
    for row_number, lead in enumerate(leads, start_row):
        values = [str(lead[key] or "") for key in KEYS]
        for column, value in enumerate(values, 1):
            cell = ws.cell(row_number, column, value)
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            cell.border = Border(top=thin, left=thin, bottom=thin, right=thin)
            cell.fill = PatternFill(
                "solid", fgColor="F5F8FA" if row_number % 2 == 0 else "FFFFFF"
            )
        profile = str(lead["profile_url"] or "")
        if profile:
            ws.cell(row_number, 3).hyperlink = profile
            ws.cell(row_number, 3).style = "Hyperlink"
        confidence = str(lead["confidence"])
        ws.cell(row_number, 13).font = Font(
            bold=True, color="1B5E20" if confidence == "高" else "8A4B08"
        )
        ws.row_dimensions[row_number].height = 82


def add_table(ws, name: str, end_row: int, style: str) -> None:
    ref = f"A{HEADER_ROW}:N{end_row}"
    table = Table(displayName=name, ref=ref)
    table.tableStyleInfo = TableStyleInfo(
        name=style,
        showFirstColumn=False,
        showLastColumn=False,
        showRowStripes=False,
        showColumnStripes=False,
    )
    ws.add_table(table)
    ws.auto_filter.ref = ref


def find_main_sheet(wb):
    for name in wb.sheetnames:
        if name.startswith("全部") or name == "潜客名单":
            return wb[name]
    return wb[wb.sheetnames[0]]


def read_handles(ws) -> list[str]:
    return [
        canonical_handle(str(ws.cell(row, 1).value or ""))
        for row in range(DATA_ROW, ws.max_row + 1)
        if str(ws.cell(row, 1).value or "").strip()
    ]


def create_workbook(leads: list[dict[str, Any]], subtitle: str) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = f"全部{len(leads)}个"
    style_title(ws, f"Instagram 假发行业潜客名单 — 全部 {len(leads)} 个", subtitle)
    style_headers(ws)
    write_rows(ws, leads, DATA_ROW)
    add_table(ws, f"LeadTable{len(leads)}", DATA_ROW + len(leads) - 1, "TableStyleMedium2")
    return wb


def append_workbook(wb: Workbook, leads: list[dict[str, Any]], subtitle: str) -> tuple[int, int]:
    ws = find_main_sheet(wb)
    existing = read_handles(ws)
    incoming = {canonical_handle(item["account"]) for item in leads}
    overlap = sorted(set(existing) & incoming)
    if overlap:
        raise ValueError("Accounts already exist: " + ", ".join("@" + item for item in overlap))
    write_rows(ws, leads, ws.max_row + 1)
    total = len(existing) + len(leads)
    ws.title = f"全部{total}个"
    ws["A1"] = f"Instagram 假发行业潜客名单 — 全部 {total} 个"
    ws["A2"] = subtitle
    if ws.tables:
        next(iter(ws.tables.values())).ref = f"A{HEADER_ROW}:N{DATA_ROW + total - 1}"
    ws.auto_filter.ref = f"A{HEADER_ROW}:N{DATA_ROW + total - 1}"

    batch_name = f"本次新增{len(leads)}个"
    if batch_name in wb.sheetnames:
        del wb[batch_name]
    batch = wb.create_sheet(batch_name, 1)
    style_title(batch, f"本次新增 {len(leads)} 个 Instagram 潜客", subtitle, "7B341E")
    style_headers(batch, "9C4221")
    write_rows(batch, leads, DATA_ROW)
    add_table(
        batch,
        f"BatchTable{total}_{len(leads)}",
        DATA_ROW + len(leads) - 1,
        "TableStyleMedium3",
    )
    return total, len(leads)


def verify_saved(path: Path, expected_total: int, expected_batch: int | None) -> None:
    wb = load_workbook(path, read_only=False, data_only=False)
    ws = find_main_sheet(wb)
    handles = read_handles(ws)
    if len(handles) != expected_total or len(set(handles)) != expected_total:
        raise RuntimeError("Saved workbook count or uniqueness check failed.")
    if ws.max_column != len(HEADERS):
        raise RuntimeError("Saved workbook column count check failed.")
    if any(ws.row_dimensions[row].hidden for row in range(DATA_ROW, ws.max_row + 1)):
        raise RuntimeError("Saved workbook contains hidden lead rows.")
    if expected_batch is not None:
        name = f"本次新增{expected_batch}个"
        if name not in wb.sheetnames or len(read_handles(wb[name])) != expected_batch:
            raise RuntimeError("Newest-batch worksheet verification failed.")


def main() -> None:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--input", required=True, type=Path, help="Validated JSON lead array")
    parser.add_argument("--output", required=True, type=Path, help="Output XLSX path")
    parser.add_argument("--append", action="store_true", help="Append to an existing workbook")
    parser.add_argument(
        "--subtitle",
        default="Live 公开数据｜人工复核名单｜未执行关注、私信或营销动作",
    )
    args = parser.parse_args()
    leads = load_leads(args.input)
    args.output.parent.mkdir(parents=True, exist_ok=True)

    if args.append:
        if not args.output.exists():
            raise FileNotFoundError(f"Append target does not exist: {args.output}")
        wb = load_workbook(args.output)
        expected_total, expected_batch = append_workbook(wb, leads, args.subtitle)
    else:
        wb = create_workbook(leads, args.subtitle)
        expected_total, expected_batch = len(leads), None

    wb.save(args.output)
    verify_saved(args.output, expected_total, expected_batch)
    message = f"Saved {args.output} with {expected_total} unique leads"
    if expected_batch is not None:
        message += f"; newest batch: {expected_batch}"
    print(message)


if __name__ == "__main__":
    main()
